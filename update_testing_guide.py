import openpyxl
import json
import os

EXCEL_PATH = '/Users/macbookpro/ErwanzCode/Pentago/Testing Guide.xlsx'
JSON_PATH = '/Users/macbookpro/ErwanzCode/Pentago/static/js/testing-guide-db.json'
JS_PATH = '/Users/macbookpro/ErwanzCode/Pentago/static/js/testing-guide-db.js'

wb = openpyxl.load_workbook(EXCEL_PATH)
db_data = {}

for name in wb.sheetnames:
    sheet = wb[name]
    metadata = {}
    header_row_idx = None
    headers = []
    
    # 1. Find the header row and extract metadata
    for r in range(1, sheet.max_row + 1):
        row_vals = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]
        
        # Check if this row is the header row
        row_str_vals = [str(x).strip().lower() for x in row_vals if x is not None]
        if any(keyword in row_str_vals for keyword in ['test id', 'no.', 'vulnerability name', 'test case']):
            header_row_idx = r
            # Trim trailing None values for headers
            while row_vals and row_vals[-1] is None:
                row_vals.pop()
            headers = [str(x) if x is not None else '' for x in row_vals]
            break
        else:
            # Look for metadata (e.g. 'Client/Project Name') in this row
            for col_idx in range(len(row_vals)):
                val_here = row_vals[col_idx]
                if val_here in ['Client/Project Name', 'Methodology', 'Scope', 'Timeline', 'OverallRisk']:
                    if col_idx + 1 < len(row_vals):
                        next_val = row_vals[col_idx + 1]
                        if next_val is not None:
                            metadata[val_here] = str(next_val).strip()

    if header_row_idx is None:
        print(f"Header row not found in sheet: {name}")
        continue
        
    # 2. Extract rows
    rows = []
    for r in range(header_row_idx + 1, sheet.max_row + 1):
        row_vals = [sheet.cell(r, c).value for c in range(1, len(headers) + 1)]
        # Skip completely empty rows
        if not any(x is not None for x in row_vals):
            continue
        row_formatted = [str(x).strip() if x is not None else '' for x in row_vals]
        rows.append(row_formatted)
        
    db_data[name] = {
        'metadata': metadata,
        'headers': headers,
        'rows': rows
    }

# Write to JSON
with open(JSON_PATH, 'w', encoding='utf-8') as f:
    json.dump(db_data, f, indent=2, ensure_ascii=False)

# Write to JS
with open(JS_PATH, 'w', encoding='utf-8') as f:
    f.write("const testingGuideData = ")
    json.dump(db_data, f, indent=2, ensure_ascii=False)
    f.write(";\n\nif (typeof module !== 'undefined' && module.exports) {\n  module.exports = testingGuideData;\n}\n")

print("Successfully updated testing guide database.")
