import os
from datetime import datetime, timedelta
import re
import functools
from flask import Flask, render_template, request, jsonify, redirect, url_for, session
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
import json
import pyotp

app = Flask(__name__)
db_path = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'pentago.db')
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{db_path}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'pentago-secret-key-12345'
db = SQLAlchemy(app)

with app.app_context():
    try:
        with db.engine.connect() as conn:
            from sqlalchemy import text
            columns = [row[1] for row in conn.execute(text("PRAGMA table_info(project)")).fetchall()]
            if 'pentest_activity' not in columns:
                conn.execute(text("ALTER TABLE project ADD COLUMN pentest_activity VARCHAR(50) DEFAULT 'Not Started'"))
            if 'retest_activity' not in columns:
                conn.execute(text("ALTER TABLE project ADD COLUMN retest_activity VARCHAR(50) DEFAULT 'Not Started'"))
            conn.commit()
    except Exception as e:
        print("Error adding columns during init:", e)

# Database Encryption Setup
from cryptography.fernet import Fernet
KEY_FILE = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'encryption.key')
if not os.path.exists(KEY_FILE):
    env_key = os.environ.get('ENCRYPTION_KEY')
    if env_key:
        try:
            cipher = Fernet(env_key.encode('utf-8'))
        except Exception:
            encryption_key = Fernet.generate_key()
            with open(KEY_FILE, 'wb') as f:
                f.write(encryption_key)
            cipher = Fernet(encryption_key)
    else:
        encryption_key = Fernet.generate_key()
        with open(KEY_FILE, 'wb') as f:
            f.write(encryption_key)
        cipher = Fernet(encryption_key)
else:
    with open(KEY_FILE, 'rb') as f:
        encryption_key = f.read()
    cipher = Fernet(encryption_key)

class EncryptedText(db.TypeDecorator):
    impl = db.Text
    cache_ok = True

    def process_bind_param(self, value, dialect):
        if value is not None:
            return cipher.encrypt(value.encode('utf-8')).decode('utf-8')
        return value

    def process_result_value(self, value, dialect):
        if value is not None:
            try:
                return cipher.decrypt(value.encode('utf-8')).decode('utf-8')
            except Exception:
                return value
        return value

# Database Models
user_client = db.Table('user_client',
    db.Column('user_id', db.Integer, db.ForeignKey('user.id', ondelete='CASCADE'), primary_key=True),
    db.Column('company_id', db.Integer, db.ForeignKey('company.id', ondelete='CASCADE'), primary_key=True)
)

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(100), unique=True, nullable=False)
    password_hash = db.Column(db.String(250), nullable=False)
    role = db.Column(db.String(50), nullable=False, default='User') # Admin, User
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    allowed_companies = db.relationship('Company', secondary=user_client, backref=db.backref('allowed_users', lazy='dynamic'))
    mfa_secret = db.Column(db.String(100), nullable=True)
    mfa_enabled = db.Column(db.Boolean, default=False)
    is_disabled = db.Column(db.Boolean, default=False)
    consultant_id = db.Column(db.Integer, db.ForeignKey('consultant.id'), nullable=True)
    consultant = db.relationship('Consultant', foreign_keys=[consultant_id])

    def to_dict(self):
        deadline_passed = False
        if not self.mfa_enabled and self.created_at:
            if (datetime.utcnow() - self.created_at) > timedelta(hours=24):
                deadline_passed = True
        return {
            'id': self.id,
            'username': self.username,
            'role': self.role,
            'created_at': self.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'allowed_companies': [c.id for c in self.allowed_companies],
            'mfa_enabled': self.mfa_enabled,
            'is_disabled': self.is_disabled or deadline_passed,
            'deadline_passed': deadline_passed,
            'consultant_id': self.consultant_id,
            'consultant_name': self.consultant.name if self.consultant else None,
            'fullname': self.consultant.name if self.consultant else self.username,
            'email': self.consultant.email if self.consultant else '',
            'member_role': self.consultant.role if self.consultant else 'Consultant'
        }

class AuditLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    username = db.Column(db.String(100), nullable=True)
    action = db.Column(db.String(100), nullable=False)
    details = db.Column(EncryptedText)
    ip_address = db.Column(db.String(50))

    def to_dict(self):
        return {
            'id': self.id,
            'timestamp': self.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            'user_id': self.user_id,
            'username': self.username or 'System',
            'action': self.action,
            'details': self.details,
            'ip_address': self.ip_address
        }

class BlockedIP(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    ip_address = db.Column(db.String(100), unique=True, nullable=False)
    reason = db.Column(db.String(250))
    blocked_at = db.Column(db.DateTime, default=datetime.utcnow)
    expires_at = db.Column(db.DateTime, nullable=True)

    def to_dict(self):
        return {
            'id': self.id,
            'ip_address': self.ip_address,
            'reason': self.reason,
            'blocked_at': self.blocked_at.strftime('%Y-%m-%d %H:%M:%S'),
            'expires_at': self.expires_at.strftime('%Y-%m-%d %H:%M:%S') if self.expires_at else 'Permanent'
        }

def log_audit(action, details):
    user_id = session.get('user_id')
    username = session.get('username')
    ip_address = request.headers.get('X-Forwarded-For', request.remote_addr)
    log_entry = AuditLog(
        user_id=user_id,
        username=username,
        action=action,
        details=details,
        ip_address=ip_address
    )
    db.session.add(log_entry)
    db.session.commit()

# --- Intrusion Detection & Rate Limiter Configuration ---

login_failures = {} # In-memory dictionary: IP -> {'attempts': count, 'locked_until': datetime}

# Basic WAF Signatures
SQLI_PATTERN = re.compile(
    r"(?i)(union\s+select|select\s+.*\s+from|insert\s+into|delete\s+from|update\s+.*\s+set|"
    r"or\s+['\"]?\d+['\"]?\s*=\s*['\"]?\d+['\"]?|admin'\s*or\s*|'or\s*1\s*=\s*1|--|\/\*|\*\/)"
)
XSS_PATTERN = re.compile(
    r"(?i)(<script|javascript:|onerror\s*=|onload\s*=|alert\(|confirm\(|prompt\(|<img|<iframe)"
)
PATH_TRAVERSAL_PATTERN = re.compile(
    r"(\.\.\/\.\.\/|\.\.\\\.\.\\|/etc/passwd|\\windows\\system32|cmd\.exe|/bin/sh|/bin/bash)"
)

def check_payload_injection(payload):
    """Recursively checks if a dictionary or string contains an injection payload."""
    if isinstance(payload, str):
        if SQLI_PATTERN.search(payload):
            return "SQL Injection Attempt"
        if XSS_PATTERN.search(payload):
            return "Cross-Site Scripting (XSS) Attempt"
        if PATH_TRAVERSAL_PATTERN.search(payload):
            return "Path Traversal/RCE Attempt"
    elif isinstance(payload, dict):
        for key, val in payload.items():
            res = check_payload_injection(val)
            if res:
                return f"{res} in field '{key}'"
    elif isinstance(payload, list):
        for item in payload:
            res = check_payload_injection(item)
            if res:
                return res
    return None

@app.before_request
def security_firewall():
    # Exclude static assets from firewall check
    if request.path.startswith('/static/'):
        return

    # Whitelist localhost
    ip_address = request.headers.get('X-Forwarded-For', request.remote_addr)
    if ip_address in ('127.0.0.1', '::1', 'localhost'):
        return

    # 1. Check if IP is blocked
    block_entry = BlockedIP.query.filter_by(ip_address=ip_address).first()
    if block_entry:
        if block_entry.expires_at and datetime.utcnow() > block_entry.expires_at:
            # Temporary block has expired, release it
            db.session.delete(block_entry)
            db.session.commit()
        else:
            return jsonify({
                'message': f'Your IP ({ip_address}) has been blocked. Reason: {block_entry.reason}'
            }), 403

    # 2. Rate Limiting for login failures
    if ip_address in login_failures:
        lock_until = login_failures[ip_address].get('locked_until')
        if lock_until and datetime.utcnow() < lock_until:
            return jsonify({
                'message': 'Too many failed login attempts. Please try again later.'
            }), 429

    # 3. WAF payload checks
    # Check query strings
    for k, v in request.args.items():
        intrusion = check_payload_injection(v)
        if intrusion:
            return trigger_waf_block(ip_address, intrusion, f"Query parameter '{k}': {v}")

    # Check JSON body payloads
    if request.is_json:
        try:
            json_data = request.get_json(silent=True)
            if json_data:
                intrusion = check_payload_injection(json_data)
                if intrusion:
                    return trigger_waf_block(ip_address, intrusion, f"JSON Payload: {json_data}")
        except Exception:
            pass

    # Check Form body payloads
    for k, v in request.form.items():
        intrusion = check_payload_injection(v)
        if intrusion:
            return trigger_waf_block(ip_address, intrusion, f"Form parameter '{k}': {v}")

def trigger_waf_block(ip_address, intrusion_type, details):
    # Log intrusion to AuditTrail
    log_entry = AuditLog(
        username="System/WAF",
        action="SECURITY_INTRUSION",
        details=f"Blocked {intrusion_type} from IP {ip_address}. Details: {details}",
        ip_address=ip_address
    )
    db.session.add(log_entry)
    
    # Auto-block the IP for 1 hour after intrusion attempt (defense-in-depth)
    expiry = datetime.utcnow() + timedelta(hours=1)
    block_entry = BlockedIP.query.filter_by(ip_address=ip_address).first()
    if not block_entry:
        block_entry = BlockedIP(
            ip_address=ip_address,
            reason=f"Automated WAF Block: {intrusion_type}",
            expires_at=expiry
        )
        db.session.add(block_entry)
    
    db.session.commit()
    
    return jsonify({
        'message': f'Request blocked by WAF. Security Intrusion Detected ({intrusion_type}). Your IP has been blocklisted.'
    }), 403

# Decorators for auth
def login_required(f):
    @functools.wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'message': 'Authentication required'}), 401
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @functools.wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'message': 'Authentication required'}), 401
        if session.get('role') != 'Admin':
            return jsonify({'message': 'Admin privileges required'}), 403
        return f(*args, **kwargs)
    return decorated_function

# Database Models
class Company(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(150), nullable=False)
    industry = db.Column(db.String(100))
    sales_name = db.Column(db.String(150))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    projects = db.relationship('Project', backref='company', lazy=True, cascade="all, delete-orphan")

    def to_dict(self):
        # Calculate overall risk metrics
        total_projects = len(self.projects)
        active_findings_count = 0
        total_risk_score = 0
        max_severity = 'Info'
        severity_map = {'Info': 0, 'Low': 1, 'Medium': 2, 'High': 3, 'Critical': 4}
        rev_severity_map = {0: 'Info', 1: 'Low', 2: 'Medium', 3: 'High', 4: 'Critical'}
        highest_val = 0

        for project in self.projects:
            proj_data = project.get_metrics()
            active_findings_count += proj_data['active_findings']
            total_risk_score += proj_data['risk_score']
            proj_highest = severity_map.get(proj_data['max_severity'], 0)
            if proj_highest > highest_val:
                highest_val = proj_highest

        max_severity = rev_severity_map[highest_val]

        return {
            'id': self.id,
            'name': self.name,
            'industry': self.industry,
            'sales_name': self.sales_name,
            'created_at': self.created_at.strftime('%Y-%m-%d'),
            'total_projects': total_projects,
            'active_findings': active_findings_count,
            'overall_risk_score': total_risk_score,
            'overall_max_severity': max_severity
        }

class Consultant(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    role = db.Column(db.String(100))
    email = db.Column(db.String(100))

    def to_dict(self):
        return {
            'id': self.id,
            'name': self.name,
            'role': self.role,
            'email': self.email
        }

class Project(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    company_id = db.Column(db.Integer, db.ForeignKey('company.id'), nullable=False)
    name = db.Column(db.String(150), nullable=False)
    po_number = db.Column(db.String(100))
    project_type = db.Column(db.String(100), default='Project Based')
    status = db.Column(db.String(50), default='In Progress') # In Progress, Completed, Retest Pending, Retest Completed
    start_date = db.Column(db.String(50))
    end_date = db.Column(db.String(50))
    description = db.Column(EncryptedText)
    summary = db.Column(EncryptedText)
    appendix = db.Column(EncryptedText)
    methodology = db.Column(db.String(50), default='Blackbox') # Blackbox, Greybox, Whitebox
    pentest_consultant_id = db.Column(db.Integer, db.ForeignKey('consultant.id'))
    retest_consultant_id = db.Column(db.Integer, db.ForeignKey('consultant.id'))
    project_manager_id = db.Column(db.Integer, db.ForeignKey('consultant.id'))
    sales_id = db.Column(db.Integer, db.ForeignKey('consultant.id'))
    pentest_consultant = db.relationship('Consultant', foreign_keys=[pentest_consultant_id])
    retest_consultant = db.relationship('Consultant', foreign_keys=[retest_consultant_id])
    project_manager = db.relationship('Consultant', foreign_keys=[project_manager_id])
    sales = db.relationship('Consultant', foreign_keys=[sales_id])
    scope = db.Column(EncryptedText)
    out_of_scope = db.Column(EncryptedText)
    access_info = db.Column(EncryptedText)
    location_type = db.Column(db.String(50), default='Remote')
    used_tools = db.Column(EncryptedText)
    threat_model = db.Column(EncryptedText)
    pentest_activity = db.Column(db.String(50), default='Not Started')
    retest_activity = db.Column(db.String(50), default='Not Started')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    findings = db.relationship('Finding', backref='project', lazy=True, cascade="all, delete-orphan")

    def get_metrics(self):
        active_findings = [f for f in self.findings if f.status == 'Open']
        risk_score = 0
        max_severity = 'Info'
        severity_map = {'Info': 0, 'Low': 1, 'Medium': 2, 'High': 3, 'Critical': 4}
        rev_severity_map = {0: 'Info', 1: 'Low', 2: 'Medium', 3: 'High', 4: 'Critical'}
        highest_val = 0
        
        severity_counts = {'Critical': 0, 'High': 0, 'Medium': 0, 'Low': 0, 'Info': 0}

        for f in active_findings:
            if f.severity in severity_counts:
                severity_counts[f.severity] += 1
                
            # Score weight
            if f.severity == 'Critical':
                risk_score += 10
            elif f.severity == 'High':
                risk_score += 7
            elif f.severity == 'Medium':
                risk_score += 4
            elif f.severity == 'Low':
                risk_score += 1
            
            val = severity_map.get(f.severity, 0)
            if val > highest_val:
                highest_val = val

        max_severity = rev_severity_map[highest_val]

        return {
            'total_findings': len(self.findings),
            'active_findings': len(active_findings),
            'risk_score': risk_score,
            'max_severity': max_severity,
            'severity_counts': severity_counts
        }

    def to_dict(self):
        metrics = self.get_metrics()
        return {
            'id': self.id,
            'company_id': self.company_id,
            'company_name': self.company.name if self.company else '',
            'name': self.name,
            'status': self.status,
            'start_date': self.start_date,
            'end_date': self.end_date,
            'description': self.description,
            'summary': self.summary,
            'appendix': self.appendix,
            'methodology': self.methodology,
            'po_number': self.po_number,
            'project_type': self.project_type,
            'pentest_consultant_id': self.pentest_consultant_id,
            'retest_consultant_id': self.retest_consultant_id,
            'project_manager_id': self.project_manager_id,
            'sales_id': self.sales_id,
            'pentest_consultant_name': self.pentest_consultant.name if self.pentest_consultant else '-',
            'retest_consultant_name': self.retest_consultant.name if self.retest_consultant else '-',
            'project_manager_name': self.project_manager.name if self.project_manager else '-',
            'sales_name': self.sales.name if self.sales else '-',
            'scope': self.scope,
            'out_of_scope': self.out_of_scope,
            'access_info': self.access_info,
            'location_type': self.location_type,
            'used_tools': self.used_tools,
            'threat_model': self.threat_model,
            'pentest_activity': self.pentest_activity,
            'retest_activity': self.retest_activity,
            'created_at': self.created_at.strftime('%Y-%m-%d'),
            'total_findings': metrics['total_findings'],
            'active_findings': metrics['active_findings'],
            'risk_score': metrics['risk_score'],
            'max_severity': metrics['max_severity'],
            'severity_counts': metrics['severity_counts']
        }

class Finding(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'), nullable=False)
    title = db.Column(EncryptedText, nullable=False)
    affected_system = db.Column(db.String(250))
    description = db.Column(EncryptedText)
    poc = db.Column(EncryptedText)
    poc_image = db.Column(db.String(500))
    poc_image_align = db.Column(db.String(50), default='center')
    poc_image_caption = db.Column(db.String(250), default='')
    exploitation = db.Column(EncryptedText)
    impact = db.Column(EncryptedText)
    script_payload = db.Column(EncryptedText)
    solution = db.Column(EncryptedText)
    reference = db.Column(EncryptedText)
    step_reproduce = db.Column(EncryptedText)
    cwe = db.Column(db.String(100))
    mitre_attack = db.Column(db.String(150))
    iso_27001 = db.Column(db.String(150))
    nist_control = db.Column(db.String(150))
    ptes_phase = db.Column(db.String(100))
    cvss_version = db.Column(db.String(10), default='v3.1') # v3.1 or v4.0
    cvss_vector = db.Column(db.String(250))
    cvss_score = db.Column(db.Float, default=0.0)
    severity = db.Column(db.String(50), default='Info') # Critical, High, Medium, Low, Info
    status = db.Column(db.String(50), default='Open') # Open, Fixed, Risk Accepted
    finding_status = db.Column(db.String(50), default='Open') # Open, Closed
    retest_evidence = db.Column(EncryptedText)
    custom_fields = db.Column(EncryptedText)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def to_dict(self):
        return {
            'id': self.id,
            'project_id': self.project_id,
            'project_name': self.project.name if self.project else '',
            'company_name': self.project.company.name if self.project and self.project.company else '',
            'title': self.title,
            'affected_system': self.affected_system,
            'description': self.description,
            'poc': self.poc,
            'poc_image': self.poc_image,
            'poc_image_align': self.poc_image_align,
            'poc_image_caption': self.poc_image_caption,
            'exploitation': self.exploitation,
            'impact': self.impact,
            'script_payload': self.script_payload,
            'solution': self.solution,
            'reference': self.reference,
            'step_reproduce': self.step_reproduce,
            'cwe': self.cwe,
            'mitre_attack': self.mitre_attack,
            'iso_27001': self.iso_27001,
            'nist_control': self.nist_control,
            'ptes_phase': self.ptes_phase,
            'cvss_version': self.cvss_version,
            'cvss_vector': self.cvss_vector,
            'cvss_score': self.cvss_score,
            'severity': self.severity,
            'status': self.status,
            'finding_status': self.finding_status,
            'retest_evidence': self.retest_evidence,
            'custom_fields': self.custom_fields,
            'created_at': self.created_at.strftime('%Y-%m-%d')
        }

class FindingTemplate(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(EncryptedText, nullable=False)
    description = db.Column(EncryptedText)
    exploitation = db.Column(EncryptedText)
    impact = db.Column(EncryptedText)
    solution = db.Column(EncryptedText)
    reference = db.Column(EncryptedText)
    step_reproduce = db.Column(EncryptedText)
    cwe = db.Column(db.String(100))
    mitre_attack = db.Column(db.String(150))
    iso_27001 = db.Column(db.String(150))
    nist_control = db.Column(db.String(150))
    ptes_phase = db.Column(db.String(100))
    cvss_version = db.Column(db.String(10), default='v3.1')
    cvss_vector = db.Column(db.String(250))
    cvss_score = db.Column(db.Float, default=0.0)
    severity = db.Column(db.String(50), default='Info')
    custom_fields = db.Column(EncryptedText)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def to_dict(self):
        return {
            'id': self.id,
            'title': self.title,
            'description': self.description,
            'exploitation': self.exploitation,
            'impact': self.impact,
            'solution': self.solution,
            'reference': self.reference,
            'step_reproduce': self.step_reproduce,
            'cwe': self.cwe,
            'mitre_attack': self.mitre_attack,
            'iso_27001': self.iso_27001,
            'nist_control': self.nist_control,
            'ptes_phase': self.ptes_phase,
            'cvss_version': self.cvss_version,
            'cvss_vector': self.cvss_vector,
            'cvss_score': self.cvss_score,
            'severity': self.severity,
            'custom_fields': self.custom_fields,
            'created_at': self.created_at.strftime('%Y-%m-%d')
        }


with app.app_context():
    db.create_all()
    from sqlalchemy import text
    try:
        db.session.execute(text("ALTER TABLE project ADD COLUMN project_manager_id INTEGER REFERENCES consultant(id)"))
        db.session.commit()
    except Exception:
        db.session.rollback()
    try:
        db.session.execute(text("ALTER TABLE project ADD COLUMN sales_id INTEGER REFERENCES consultant(id)"))
        db.session.commit()
    except Exception:
        db.session.rollback()
    try:
        db.session.execute(text("ALTER TABLE finding ADD COLUMN custom_fields TEXT"))
        db.session.commit()
    except Exception:
        db.session.rollback()
    try:
        db.session.execute(text("ALTER TABLE finding_template ADD COLUMN custom_fields TEXT"))
        db.session.commit()
    except Exception:
        db.session.rollback()
    try:
        db.session.execute(text("ALTER TABLE user ADD COLUMN mfa_secret TEXT"))
        db.session.commit()
    except Exception:
        db.session.rollback()
    try:
        db.session.execute(text("ALTER TABLE user ADD COLUMN mfa_enabled BOOLEAN DEFAULT 0"))
        db.session.commit()
    except Exception:
        db.session.rollback()
    try:
        db.session.execute(text("ALTER TABLE user ADD COLUMN is_disabled BOOLEAN DEFAULT 0"))
        db.session.commit()
    except Exception:
        db.session.rollback()
    try:
        db.session.execute(text("ALTER TABLE user ADD COLUMN consultant_id INTEGER REFERENCES consultant(id)"))
        db.session.commit()
    except Exception:
        db.session.rollback()
    # Check if Users are empty, then seed default credentials
    if not User.query.first():
        admin_user = User(
            username="admin",
            password_hash=generate_password_hash("admin123"),
            role="Admin"
        )
        regular_user = User(
            username="user",
            password_hash=generate_password_hash("user123"),
            role="User"
        )
        db.session.add(admin_user)
        db.session.add(regular_user)
        db.session.commit()

    # Check if empty, then seed demo data
    if not Company.query.first():
        demo_company = Company(name="Fintech Global Corp", industry="Financial Technology")
        db.session.add(demo_company)
        db.session.commit()
        
        demo_project = Project(
            company_id=demo_company.id,
            name="External Pentest & API Assessment",
            status="In Progress",
            start_date="2026-06-01",
            end_date="2026-06-30",
            description="Comprehensive security assessment of Fintech Global external facing web applications and core APIs."
        )
        db.session.add(demo_project)
        db.session.commit()
        
        # 1. SQL Injection (Critical/High - CVSS 4.0)
        finding1 = Finding(
            project_id=demo_project.id,
            title="SQL Injection on User Authentication Endpoint",
            affected_system="https://api.fintechglobal.com/v1/auth/login",
            description="A SQL injection vulnerability exists in the login endpoint via the 'username' parameter. Unsanitized input is passed directly to the database query, allowing an attacker to manipulate queries and bypass authentication.",
            poc="POST /v1/auth/login HTTP/1.1\nHost: api.fintechglobal.com\nContent-Type: application/json\n\n{\n  \"username\": \"admin' OR '1'='1\",\n  \"password\": \"anything\"\n}",
            poc_image="https://images.unsplash.com/photo-1601597111158-2fceff270190?w=800",
            exploitation="By entering SQL payload admin' OR '1'='1 in the username field, the database query executes as SELECT * FROM users WHERE username = 'admin' OR '1'='1' AND password = '...' which evaluates to true, bypassing the password validation step and logging in as the first user (typically administrative user).",
            impact="Complete compromise of database confidentiality and integrity. An attacker can read, modify, or delete all data inside the application database, including user credentials and transaction records.",
            solution="Use parameterized queries / prepared statements for all database interactions. Implement input validation and employ a Web Application Firewall (WAF) as defense-in-depth.",
            reference="OWASP Top 10 A03:2021-Injection, PortSwigger SQL Injection Academy",
            step_reproduce="1. Send a POST request to /v1/auth/login.\n2. Supply 'admin' OR '1'='1' as username parameter.\n3. Notice that login succeeds without a valid password.",
            cwe="CWE-89: Improper Neutralization of Special Elements used in an SQL Command",
            mitre_attack="T1190: Exploit Public-Facing Application",
            iso_27001="A.12.6.1: Management of Technical Vulnerabilities",
            nist_control="RA-5: Vulnerability Monitoring and Scanning",
            ptes_phase="Exploitation",
            cvss_version="v4.0",
            cvss_vector="CVSS:4.0/AV:N/AC:L/AT:N/PR:N/UI:N/VC:H/VI:H/VA:H",
            cvss_score=9.3,
            severity="Critical",
            status="Open"
        )
        
        # 2. Reflected XSS (Medium - CVSS 4.0)
        finding2 = Finding(
            project_id=demo_project.id,
            title="Reflected Cross-Site Scripting (XSS) on Query Parameter",
            affected_system="https://fintechglobal.com/search?q=query",
            description="The search functionality fails to sanitize user input before rendering it in the HTML response page. An attacker can craft a malicious link containing script payloads that execute in the victim's browser context.",
            poc="GET /search?q=%3Cscript%3Ealert(document.cookie)%3C/script%3E HTTP/1.1\nHost: fintechglobal.com",
            poc_image="https://images.unsplash.com/photo-1526374965328-7f61d4dc18c5?w=800",
            exploitation="An attacker generates a malicious search link containing script tags and sends it to a victim. When the victim clicks the link, the script executes automatically, accessing the session cookie data.",
            impact="Session hijacking. An attacker can steal user cookies or session tokens, allowing them to impersonate the victim on the platform.",
            solution="Sanitize all user inputs and utilize contextual HTML encoding (e.g. encode for HTML body, attributes, JS contexts). Implement a strong Content Security Policy (CSP).",
            reference="OWASP Top 10 A03:2021-Injection, PortSwigger XSS Academy",
            step_reproduce="1. Navigate to the search endpoint.\n2. Input <script>alert(1)</script> in the search field.\n3. Observe the JavaScript pop-up displaying in the browser.",
            cwe="CWE-79: Improper Neutralization of Input During Web Page Generation",
            mitre_attack="T1190: Exploit Public-Facing Application",
            iso_27001="A.14.2.5: Secure System Engineering Principles",
            nist_control="SI-10: Information Input Validation",
            ptes_phase="Exploitation",
            cvss_version="v4.0",
            cvss_vector="CVSS:4.0/AV:N/AC:L/AT:N/PR:N/UI:A/VC:L/VI:L/VA:N",
            cvss_score=5.3,
            severity="Medium",
            status="Open"
        )
        
        # 3. Missing Security Headers (Low/Info - CVSS 4.0)
        finding3 = Finding(
            project_id=demo_project.id,
            title="Missing HTTP Security Headers (HSTS, CSP)",
            affected_system="https://fintechglobal.com",
            description="The web server does not return essential security headers (such as Content-Security-Policy or Strict-Transport-Security), exposing the client browser to clickjacking, protocol downgrades, and content injection.",
            poc="curl -I https://fintechglobal.com\n\nHTTP/1.1 200 OK\nContent-Type: text/html\n(Note: Absence of Strict-Transport-Security, CSP, X-Frame-Options)",
            exploitation="An attacker can perform clickjacking attacks or intercept HTTP traffic due to protocol downgrade to HTTP.",
            impact="Client browsers are unprotected against modern web client-side attacks.",
            solution="Configure the web server (Nginx/Apache) to append HTTP security headers: Content-Security-Policy, Strict-Transport-Security, and X-Content-Type-Options.",
            reference="OWASP Secure Headers Project, OWASP Top 10 A05:2021-Security Misconfiguration",
            step_reproduce="1. Execute curl request against the root URL.\n2. Verify the HTTP headers returned in the response.",
            cwe="CWE-693: Protection Mechanism Failure",
            mitre_attack="T1082: System Information Discovery",
            iso_27001="A.14.1.1: Information Security Requirements Analysis and Specification",
            nist_control="SC-8: Transmission Confidentiality and Integrity",
            ptes_phase="Vulnerability Analysis",
            cvss_version="v4.0",
            cvss_vector="CVSS:4.0/AV:N/AC:L/AT:N/PR:N/UI:N/VC:N/VI:N/VA:N",
            cvss_score=0.0,
            severity="Info",
            status="Open"
        )
        
        db.session.add(finding1)
        db.session.add(finding2)
        db.session.add(finding3)
        db.session.commit()

# Upload Handler
import uuid
UPLOAD_FOLDER = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'static', 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

@app.route('/api/upload', methods=['POST'])
@login_required
def api_upload():
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ['.png', '.jpg', '.jpeg', '.gif', '.webp']:
        return jsonify({'error': 'Only image files (.png, .jpg, .jpeg, .gif, .webp) are allowed'}), 400
    filename = f"poc_{uuid.uuid4().hex}{ext}"
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    file.save(filepath)
    file_url = f"/static/uploads/{filename}"
    log_audit('UPLOAD_FILE', f"Uploaded image file: {filename}")
    return jsonify({'url': file_url})

# Testing Guides CRUD API
GUIDE_FILE_PATH = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'static', 'js', 'testing-guide-db.json')
GUIDE_JS_PATH = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'static', 'js', 'testing-guide-db.js')

@app.route('/api/testing-guides', methods=['GET'])
@login_required
def get_testing_guides():
    data = {}
    if not os.path.exists(GUIDE_FILE_PATH):
        if os.path.exists(GUIDE_JS_PATH):
            try:
                with open(GUIDE_JS_PATH, 'r', encoding='utf-8') as f:
                    content = f.read()
                json_str = content.split('const testingGuideData = ')[1].split(';')[0].strip()
                data = json.loads(json_str)
                with open(GUIDE_FILE_PATH, 'w', encoding='utf-8') as f:
                    json.dump(data, f, indent=2, ensure_ascii=False)
            except Exception as e:
                return jsonify({'error': f'Failed to parse JS: {str(e)}'}), 500
    else:
        try:
            with open(GUIDE_FILE_PATH, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except Exception as e:
            return jsonify({'error': str(e)}), 500
            
    response = jsonify(data)
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

@app.route('/api/testing-guides', methods=['POST'])
@login_required
def save_testing_guides():
    if session.get('role') != 'Admin':
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.json
    try:
        with open(GUIDE_FILE_PATH, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        with open(GUIDE_JS_PATH, 'w', encoding='utf-8') as f:
            f.write("const testingGuideData = ")
            json.dump(data, f, indent=2, ensure_ascii=False)
            f.write(";\n\nif (typeof module !== 'undefined' && module.exports) {\n  module.exports = testingGuideData;\n}\n")
        log_audit('SAVE_TESTING_GUIDES', "Updated Testing Guides database")
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Reference Library API
REFERENCES_FILE_PATH = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'static', 'js', 'references-db.json')

@app.route('/api/references', methods=['GET'])
@login_required
def get_references():
    if not os.path.exists(REFERENCES_FILE_PATH):
        # initialize empty categories if missing
        default = {"categories": {}}
        with open(REFERENCES_FILE_PATH, 'w', encoding='utf-8') as f:
            json.dump(default, f, indent=2, ensure_ascii=False)
    try:
        with open(REFERENCES_FILE_PATH, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    return jsonify(data)

@app.route('/api/references', methods=['POST'])
@login_required
def save_references():
    if session.get('role') != 'Admin' and session.get('role') != 'User':
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.json
    try:
        with open(REFERENCES_FILE_PATH, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        log_audit('SAVE_REFERENCES', "Updated Reference Library database")
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# Testing Guides Excel Template Download
@app.route('/api/testing-guides/template', methods=['GET'])
@login_required
def download_testing_guide_template():
    if session.get('role') != 'Admin':
        return jsonify({'error': 'Unauthorized'}), 403
    from openpyxl import Workbook
    import io
    from flask import send_file
    category = request.args.get('category', '')
    if not category:
        return jsonify({'error': 'Category parameter is required'}), 400
    if not os.path.exists(GUIDE_FILE_PATH):
        return jsonify({'error': 'Testing guides database not found'}), 404
    with open(GUIDE_FILE_PATH, 'r', encoding='utf-8') as f:
        guide_data = json.load(f)
    if category not in guide_data:
        return jsonify({'error': 'Category not found'}), 404
    headers = guide_data[category].get('headers', [
        "Test ID", "Area / Platform", "Test Case", "Priority", 
        "Purpose", "Steps", "Expected Result", "Tools", "CWE"
    ])
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    ws.append(headers)
    if len(guide_data[category].get('rows', [])) > 0:
        ws.append(guide_data[category]['rows'][0])
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return send_file(
        file_stream,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"template_{category.lower().replace(' ', '_')}.xlsx"
    )

# Testing Guides Excel Import
@app.route('/api/testing-guides/import', methods=['POST'])
@login_required
def import_testing_guide_excel():
    if session.get('role') != 'Admin':
        return jsonify({'error': 'Unauthorized'}), 403
    from openpyxl import load_workbook
    import io
    category = request.form.get('category', '')
    if not category:
        return jsonify({'error': 'Category is required'}), 400
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Empty filename'}), 400
    try:
        wb = load_workbook(io.BytesIO(file.read()))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return jsonify({'error': 'Excel file is empty'}), 400
        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        data_rows = []
        for r in rows[1:]:
            if all(cell is None or str(cell).strip() == "" for cell in r):
                continue
            data_rows.append([str(cell) if cell is not None else "" for cell in r])
        if not os.path.exists(GUIDE_FILE_PATH):
            return jsonify({'error': 'Testing guides database not found'}), 404
        with open(GUIDE_FILE_PATH, 'r', encoding='utf-8') as f:
            guide_data = json.load(f)
        if category not in guide_data:
            guide_data[category] = {
                "metadata": {},
                "headers": headers,
                "rows": []
            }
        guide_data[category]["rows"] = data_rows
        guide_data[category]["headers"] = headers
        with open(GUIDE_FILE_PATH, 'w', encoding='utf-8') as f:
            json.dump(guide_data, f, indent=2, ensure_ascii=False)
        with open(GUIDE_JS_PATH, 'w', encoding='utf-8') as f:
            f.write("const testingGuideData = ")
            json.dump(guide_data, f, indent=2, ensure_ascii=False)
            f.write(";\n\nif (typeof module !== 'undefined' && module.exports) {\n  module.exports = testingGuideData;\n}\n")
        log_audit('IMPORT_TESTING_GUIDES', f"Imported {len(data_rows)} rows for guide category '{category}'")
        return jsonify({'success': True, 'count': len(data_rows)})
    except Exception as e:
        return jsonify({'error': f'Failed to process Excel file: {str(e)}'}), 500

# Frameworks Reference API
FRAMEWORKS_FILE_PATH = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'static', 'js', 'frameworks-db.json')

@app.route('/api/frameworks', methods=['GET'])
@login_required
def get_frameworks():
    try:
        if not os.path.exists(FRAMEWORKS_FILE_PATH):
            return jsonify({'error': 'Frameworks database not found'}), 404
        with open(FRAMEWORKS_FILE_PATH, 'r', encoding='utf-8') as f:
            data = json.load(f)
        response = jsonify(data)
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        return response
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/frameworks', methods=['POST'])
@login_required
def save_frameworks():
    if session.get('role') != 'Admin':
        return jsonify({'error': 'Unauthorized'}), 403
    data = request.json
    try:
        with open(FRAMEWORKS_FILE_PATH, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        log_audit('SAVE_FRAMEWORKS', "Updated Security Frameworks database")
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Frameworks Excel Template Download
@app.route('/api/frameworks/template', methods=['GET'])
@login_required
def download_framework_template():
    if session.get('role') != 'Admin':
        return jsonify({'error': 'Unauthorized'}), 403
    from openpyxl import Workbook
    import io
    from flask import send_file
    category = request.args.get('category', '')
    if not category:
        return jsonify({'error': 'Category parameter is required'}), 400
    if not os.path.exists(FRAMEWORKS_FILE_PATH):
        return jsonify({'error': 'Frameworks database not found'}), 404
    with open(FRAMEWORKS_FILE_PATH, 'r', encoding='utf-8') as f:
        framework_data = json.load(f)
    if category not in framework_data:
        return jsonify({'error': 'Framework not found'}), 404
    headers = framework_data[category].get('headers', ["ID", "Name", "Description"])
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    ws.append(headers)
    if len(framework_data[category].get('rows', [])) > 0:
        ws.append(framework_data[category]['rows'][0])
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    return send_file(
        file_stream,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"template_{category.lower().replace(' ', '_')}.xlsx"
    )

# Frameworks Excel Import
@app.route('/api/frameworks/import', methods=['POST'])
@login_required
def import_framework_excel():
    if session.get('role') != 'Admin':
        return jsonify({'error': 'Unauthorized'}), 403
    from openpyxl import load_workbook
    import io
    category = request.form.get('category', '')
    if not category:
        return jsonify({'error': 'Category/Framework name is required'}), 400
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Empty filename'}), 400
    try:
        wb = load_workbook(io.BytesIO(file.read()))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return jsonify({'error': 'Excel file is empty'}), 400
        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        data_rows = []
        for r in rows[1:]:
            if all(cell is None or str(cell).strip() == "" for cell in r):
                continue
            data_rows.append([str(cell) if cell is not None else "" for cell in r])
        if not os.path.exists(FRAMEWORKS_FILE_PATH):
            return jsonify({'error': 'Frameworks database not found'}), 404
        with open(FRAMEWORKS_FILE_PATH, 'r', encoding='utf-8') as f:
            framework_data = json.load(f)
        if category not in framework_data:
            framework_data[category] = {
                "description": f"Imported reference guide for {category}",
                "headers": headers,
                "rows": []
            }
        framework_data[category]["rows"] = data_rows
        framework_data[category]["headers"] = headers
        with open(FRAMEWORKS_FILE_PATH, 'w', encoding='utf-8') as f:
            json.dump(framework_data, f, indent=2, ensure_ascii=False)
        log_audit('IMPORT_FRAMEWORKS', f"Imported {len(data_rows)} rows for framework category '{category}'")
        return jsonify({'success': True, 'count': len(data_rows)})
    except Exception as e:
        return jsonify({'error': f'Failed to process Excel file: {str(e)}'}), 500

# Web Routes
@app.route('/')
def index():
    return render_template('index.html')

# API Routes
@app.route('/api/dashboard', methods=['GET'])
@login_required
def get_dashboard_stats():
    user = User.query.get(session['user_id'])
    if user.role != 'Admin':
        allowed_company_ids = [c.id for c in user.allowed_companies]
        companies = Company.query.filter(Company.id.in_(allowed_company_ids)).all()
        projects = Project.query.filter(Project.company_id.in_(allowed_company_ids)).all()
        allowed_project_ids = [p.id for p in projects]
        findings = Finding.query.filter(Finding.project_id.in_(allowed_project_ids)).all()
    else:
        companies = Company.query.all()
        projects = Project.query.all()
        findings = Finding.query.all()
    consultants = Consultant.query.order_by(Consultant.name).all()

    # Collect all available years
    years_set = set()
    for p in projects:
        if p.start_date and len(p.start_date) >= 4:
            years_set.add(p.start_date[:4])
        elif p.created_at:
            years_set.add(p.created_at.strftime('%Y'))
    for c in companies:
        if c.created_at:
            years_set.add(c.created_at.strftime('%Y'))
            
    # Add current year if empty
    if not years_set:
        years_set.add(datetime.utcnow().strftime('%Y'))
        
    available_years = sorted(list(years_set), reverse=True)

    selected_year = request.args.get('year', 'all')
    if selected_year and selected_year != 'all':
        projects = [p for p in projects if (p.start_date and p.start_date.startswith(selected_year)) or (not p.start_date and p.created_at and p.created_at.strftime('%Y') == selected_year)]
        # Filter companies that were created in that year OR have projects in that year
        companies = [c for c in companies if (c.created_at and c.created_at.strftime('%Y') == selected_year) or any((p.start_date and p.start_date.startswith(selected_year)) or (not p.start_date and p.created_at and p.created_at.strftime('%Y') == selected_year) for p in c.projects)]
        project_ids = [p.id for p in projects]
        findings = [f for f in findings if f.project_id in project_ids]

    total_companies = len(companies)
    total_projects = len(projects)
    total_findings = len(findings)
    open_findings = len([f for f in findings if f.status == 'Open'])

    # Get status breakdown
    proj_status = {'In Progress': 0, 'Completed': 0, 'Retest Pending': 0, 'Retest Completed': 0}
    for p in projects:
        proj_status[p.status] = proj_status.get(p.status, 0) + 1

    # Severity distribution
    severity_dist = {'Critical': 0, 'High': 0, 'Medium': 0, 'Low': 0, 'Info': 0}
    for f in findings:
        if f.status == 'Open':
            severity_dist[f.severity] = severity_dist.get(f.severity, 0) + 1

    # Consultants project progress
    consultants_progress = []
    for c in consultants:
        role_lower = (c.role or '').lower()
        is_pm = 'pm' in role_lower or 'project manager' in role_lower or 'manager' in role_lower or 'coordinator' in role_lower
        is_sales = 'sales' in role_lower or 'marketing' in role_lower or 'account' in role_lower or 'bd' in role_lower or 'business development' in role_lower

        if is_pm:
            project_filter = (Project.project_manager_id == c.id)
        elif is_sales:
            project_filter = (Project.sales_id == c.id)
        else:
            project_filter = ((Project.pentest_consultant_id == c.id) | (Project.retest_consultant_id == c.id))

        if user.role != 'Admin':
            c_projects = Project.query.filter(
                (Project.company_id.in_(allowed_company_ids)) & project_filter
            ).all()
        else:
            c_projects = Project.query.filter(project_filter).all()
        
        # Apply year filter to consultant projects
        if selected_year and selected_year != 'all':
            c_projects = [p for p in c_projects if (p.start_date and p.start_date.startswith(selected_year)) or (not p.start_date and p.created_at and p.created_at.strftime('%Y') == selected_year)]
            
        status_counts = {'In Progress': 0, 'Completed': 0, 'Retest Pending': 0, 'Retest Completed': 0}
        for p in c_projects:
            if p.status in status_counts:
                status_counts[p.status] += 1
                
        consultants_progress.append({
            'id': c.id,
            'name': c.name,
            'role': c.role or 'Cybersecurity Consultant',
            'email': c.email or '-',
            'total_projects': len(c_projects),
            'status_counts': status_counts
        })

    return jsonify({
        'total_companies': total_companies,
        'total_projects': total_projects,
        'total_findings': total_findings,
        'open_findings': open_findings,
        'project_statuses': proj_status,
        'severity_distribution': severity_dist,
        'consultants_progress': consultants_progress,
        'available_years': available_years
    })

# Company CRUD
@app.route('/api/companies', methods=['GET', 'POST'])
@login_required
def api_companies():
    user = User.query.get(session['user_id'])
    if request.method == 'GET':
        if user.role != 'Admin':
            companies = user.allowed_companies
        else:
            companies = Company.query.order_by(Company.name).all()
        return jsonify([c.to_dict() for c in companies])
    elif request.method == 'POST':
        if user.role != 'Admin':
            return jsonify({'message': 'Only Admins can create clients'}), 403
        data = request.json
        company = Company(name=data['name'], industry=data.get('industry', ''), sales_name=data.get('sales_name', ''))
        year = data.get('year')
        if year:
            try:
                now = datetime.utcnow()
                company.created_at = datetime(int(year), now.month, now.day, now.hour, now.minute, now.second)
            except ValueError:
                pass
        db.session.add(company)
        db.session.commit()
        log_audit('CREATE_COMPANY', f"Created company: {company.name}")
        return jsonify(company.to_dict()), 201

@app.route('/api/companies/<int:company_id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def api_company(company_id):
    user = User.query.get(session['user_id'])
    company = Company.query.get_or_404(company_id)
    if user.role != 'Admin' and company_id not in [c.id for c in user.allowed_companies]:
        return jsonify({'message': 'Unauthorized'}), 403

    if request.method == 'GET':
        return jsonify(company.to_dict())
    elif request.method == 'PUT':
        if user.role != 'Admin':
            return jsonify({'message': 'Only Admins can modify clients'}), 403
        data = request.json
        company.name = data['name']
        company.industry = data.get('industry', company.industry)
        company.sales_name = data.get('sales_name', company.sales_name)
        year = data.get('year')
        if year:
            try:
                curr = company.created_at or datetime.utcnow()
                try:
                    company.created_at = datetime(int(year), curr.month, curr.day, curr.hour, curr.minute, curr.second)
                except ValueError:
                    company.created_at = datetime(int(year), 1, 1, 12, 0, 0)
            except ValueError:
                pass
        db.session.commit()
        log_audit('UPDATE_COMPANY', f"Updated company: {company.name}")
        return jsonify(company.to_dict())
    elif request.method == 'DELETE':
        if user.role != 'Admin':
            return jsonify({'message': 'Only Admins can delete clients'}), 403
        comp_name = company.name
        db.session.delete(company)
        db.session.commit()
        log_audit('DELETE_COMPANY', f"Deleted company: {comp_name}")
        return jsonify({'message': 'Company deleted successfully'})

def check_edit_project_permission(user, project):
    if user.role == 'Admin':
        return True
    
    # Query all consultants and find a match in memory
    consultant = None
    consultants = Consultant.query.all()
    username_lower = user.username.lower()
    for c in consultants:
        if c.name.lower() == username_lower:
            consultant = c
            break
        if c.email:
            email_prefix = c.email.split('@')[0].lower()
            if email_prefix == username_lower:
                consultant = c
                break
                
    if user.consultant_id:
        consultant = user.consultant

    if consultant:
        role_lower = (consultant.role or '').lower()
        if 'leader' in role_lower or 'lead' in role_lower:
            return True
        if project.pentest_consultant_id and consultant.id == project.pentest_consultant_id:
            return True
        if project.retest_consultant_id and consultant.id == project.retest_consultant_id:
            return True
        if project.project_manager_id and consultant.id == project.project_manager_id:
            return True
        if project.sales_id and consultant.id == project.sales_id:
            return True
            
    return False

# Project CRUD
@app.route('/api/projects', methods=['GET', 'POST'])
@login_required
def api_projects():
    user = User.query.get(session['user_id'])
    if request.method == 'GET':
        company_id = request.args.get('company_id', type=int)
        from sqlalchemy.orm import defer
        deferred_cols = (
            defer(Project.description), defer(Project.summary), defer(Project.appendix),
            defer(Project.scope), defer(Project.out_of_scope), defer(Project.access_info),
            defer(Project.used_tools), defer(Project.threat_model)
        )
        if company_id:
            if user.role != 'Admin' and company_id not in [c.id for c in user.allowed_companies]:
                return jsonify({'message': 'Unauthorized'}), 403
            projects = Project.query.options(*deferred_cols).filter_by(company_id=company_id).order_by(Project.created_at.desc()).all()
        else:
            if user.role != 'Admin':
                allowed_ids = [c.id for c in user.allowed_companies]
                projects = Project.query.options(*deferred_cols).filter(Project.company_id.in_(allowed_ids)).order_by(Project.created_at.desc()).all()
            else:
                projects = Project.query.options(*deferred_cols).order_by(Project.created_at.desc()).all()
        return jsonify([p.to_dict() for p in projects])
    elif request.method == 'POST':
        # Check if user is Team Leader or Admin
        is_allowed_to_create = False
        if user.role == 'Admin':
            is_allowed_to_create = True
        else:
            consultants = Consultant.query.all()
            username_lower = user.username.lower()
            for c in consultants:
                if c.name.lower() == username_lower or (c.email and c.email.split('@')[0].lower() == username_lower):
                    role_lower = (c.role or '').lower()
                    if 'leader' in role_lower or 'lead' in role_lower:
                        is_allowed_to_create = True
                        break
        if not is_allowed_to_create:
            return jsonify({'message': 'Unauthorized: Only Team Leaders or Admins can create new projects.'}), 403

        data = request.json
        company_id = data.get('company_id')
        if user.role != 'Admin' and company_id not in [c.id for c in user.allowed_companies]:
            return jsonify({'message': 'Unauthorized to create project for this client'}), 403
        project = Project(
            company_id=company_id,
            name=data['name'],
            status=data.get('status', 'In Progress'),
            start_date=data.get('start_date', ''),
            end_date=data.get('end_date', ''),
            description=data.get('description', ''),
            summary=data.get('summary', ''),
            appendix=data.get('appendix', ''),
            methodology=data.get('methodology', 'Blackbox'),
            po_number=data.get('po_number', ''),
            project_type=data.get('project_type', 'Project Based'),
            pentest_consultant_id=data.get('pentest_consultant_id'),
            retest_consultant_id=data.get('retest_consultant_id'),
            project_manager_id=data.get('project_manager_id'),
            sales_id=data.get('sales_id'),
            scope=data.get('scope', ''),
            out_of_scope=data.get('out_of_scope', ''),
            access_info=data.get('access_info', ''),
            location_type=data.get('location_type', 'Remote'),
            used_tools=data.get('used_tools', ''),
            threat_model=data.get('threat_model', '')
        )
        db.session.add(project)
        db.session.commit()
        log_audit('CREATE_PROJECT', f"Created project: {project.name} (Company: {project.company.name if project.company else ''})")
        return jsonify(project.to_dict()), 201

@app.route('/api/projects/<int:project_id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def api_project(project_id):
    user = User.query.get(session['user_id'])
    project = Project.query.get_or_404(project_id)
    if user.role != 'Admin' and project.company_id not in [c.id for c in user.allowed_companies]:
        return jsonify({'message': 'Unauthorized'}), 403

    if request.method in ['PUT', 'DELETE']:
        if not check_edit_project_permission(user, project):
            return jsonify({'message': 'Unauthorized: Only assigned Pentest Consultant, Team Leader, or Admin can edit/delete this project.'}), 403

    if request.method == 'GET':
        return jsonify(project.to_dict())
    elif request.method == 'PUT':
        data = request.json
        project.name = data.get('name', project.name)
        project.status = data.get('status', project.status)
        project.start_date = data.get('start_date', project.start_date)
        project.end_date = data.get('end_date', project.end_date)
        project.description = data.get('description', project.description)
        project.summary = data.get('summary', project.summary)
        project.appendix = data.get('appendix', project.appendix)
        project.methodology = data.get('methodology', project.methodology)
        project.po_number = data.get('po_number', project.po_number)
        project.project_type = data.get('project_type', project.project_type)
        project.pentest_consultant_id = data.get('pentest_consultant_id', project.pentest_consultant_id)
        project.retest_consultant_id = data.get('retest_consultant_id', project.retest_consultant_id)
        project.project_manager_id = data.get('project_manager_id', project.project_manager_id)
        project.sales_id = data.get('sales_id', project.sales_id)
        project.scope = data.get('scope', project.scope)
        project.out_of_scope = data.get('out_of_scope', project.out_of_scope)
        project.access_info = data.get('access_info', project.access_info)
        project.location_type = data.get('location_type', project.location_type)
        project.used_tools = data.get('used_tools', project.used_tools)
        project.threat_model = data.get('threat_model', project.threat_model)
        project.pentest_activity = data.get('pentest_activity', project.pentest_activity)
        project.retest_activity = data.get('retest_activity', project.retest_activity)
        db.session.commit()
        log_audit('UPDATE_PROJECT', f"Updated project: {project.name}")
        return jsonify(project.to_dict())
    elif request.method == 'DELETE':
        proj_name = project.name
        db.session.delete(project)
        db.session.commit()
        log_audit('DELETE_PROJECT', f"Deleted project: {proj_name}")
        return jsonify({'message': 'Project deleted successfully'})

# Finding CRUD
@app.route('/api/findings', methods=['GET', 'POST'])
@login_required
def api_findings():
    user = User.query.get(session['user_id'])
    if request.method == 'GET':
        project_id = request.args.get('project_id', type=int)
        if project_id:
            project = Project.query.get_or_404(project_id)
            if user.role != 'Admin' and project.company_id not in [c.id for c in user.allowed_companies]:
                return jsonify({'message': 'Unauthorized'}), 403
            findings = Finding.query.filter_by(project_id=project_id).order_by(Finding.cvss_score.desc()).all()
        else:
            if user.role != 'Admin':
                allowed_company_ids = [c.id for c in user.allowed_companies]
                projects = Project.query.filter(Project.company_id.in_(allowed_company_ids)).all()
                allowed_project_ids = [p.id for p in projects]
                findings = Finding.query.filter(Finding.project_id.in_(allowed_project_ids)).order_by(Finding.cvss_score.desc()).all()
            else:
                findings = Finding.query.order_by(Finding.cvss_score.desc()).all()
        return jsonify([f.to_dict() for f in findings])
    elif request.method == 'POST':
        data = request.json
        project_id = data.get('project_id')
        project = Project.query.get_or_404(project_id)
        if user.role != 'Admin' and project.company_id not in [c.id for c in user.allowed_companies]:
            return jsonify({'message': 'Unauthorized to add finding for this client'}), 403
        if not check_edit_project_permission(user, project):
            return jsonify({'message': 'Unauthorized: Only assigned Pentest Consultant, Team Leader, or Admin can add findings to this project.'}), 403
        finding = Finding(
            project_id=project_id,
            title=data['title'],
            affected_system=data.get('affected_system', ''),
            description=data.get('description', ''),
            poc=data.get('poc', ''),
            poc_image=data.get('poc_image', ''),
            poc_image_align=data.get('poc_image_align', 'center'),
            poc_image_caption=data.get('poc_image_caption', ''),
            exploitation=data.get('exploitation', ''),
            impact=data.get('impact', ''),
            script_payload=data.get('script_payload', ''),
            solution=data.get('solution', ''),
            reference=data.get('reference', ''),
            step_reproduce=data.get('step_reproduce', ''),
            cwe=data.get('cwe', ''),
            mitre_attack=data.get('mitre_attack', ''),
            iso_27001=data.get('iso_27001', ''),
            nist_control=data.get('nist_control', ''),
            ptes_phase=data.get('ptes_phase', ''),
            cvss_version=data.get('cvss_version', 'v3.1'),
            cvss_vector=data.get('cvss_vector', ''),
            cvss_score=float(data.get('cvss_score', 0.0)),
            severity=data.get('severity', 'Info'),
            status=data.get('status', 'Open'),
            finding_status=data.get('finding_status', 'Open'),
            retest_evidence=data.get('retest_evidence', ''),
            custom_fields=data.get('custom_fields', '')
        )
        db.session.add(finding)
        db.session.commit()
        log_audit('CREATE_FINDING', f"Created finding: {finding.title} (Project: {finding.project.name if finding.project else ''})")
        return jsonify(finding.to_dict()), 201

@app.route('/api/findings/<int:finding_id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def api_finding(finding_id):
    user = User.query.get(session['user_id'])
    finding = Finding.query.get_or_404(finding_id)
    if user.role != 'Admin' and finding.project.company_id not in [c.id for c in user.allowed_companies]:
        return jsonify({'message': 'Unauthorized'}), 403

    if request.method in ['PUT', 'DELETE']:
        if not check_edit_project_permission(user, finding.project):
            return jsonify({'message': 'Unauthorized: Only assigned Pentest Consultant, Team Leader, or Admin can edit/delete findings.'}), 403

    if request.method == 'GET':
        return jsonify(finding.to_dict())
    elif request.method == 'PUT':
        data = request.json
        finding.title = data['title']
        finding.affected_system = data.get('affected_system', finding.affected_system)
        finding.description = data.get('description', finding.description)
        finding.poc = data.get('poc', finding.poc)
        finding.poc_image = data.get('poc_image', finding.poc_image)
        finding.poc_image_align = data.get('poc_image_align', finding.poc_image_align)
        finding.poc_image_caption = data.get('poc_image_caption', finding.poc_image_caption)
        finding.exploitation = data.get('exploitation', finding.exploitation)
        finding.impact = data.get('impact', finding.impact)
        finding.script_payload = data.get('script_payload', finding.script_payload)
        finding.solution = data.get('solution', finding.solution)
        finding.reference = data.get('reference', finding.reference)
        finding.step_reproduce = data.get('step_reproduce', finding.step_reproduce)
        finding.cwe = data.get('cwe', finding.cwe)
        finding.mitre_attack = data.get('mitre_attack', finding.mitre_attack)
        finding.iso_27001 = data.get('iso_27001', finding.iso_27001)
        finding.nist_control = data.get('nist_control', finding.nist_control)
        finding.ptes_phase = data.get('ptes_phase', finding.ptes_phase)
        finding.cvss_version = data.get('cvss_version', finding.cvss_version)
        finding.cvss_vector = data.get('cvss_vector', finding.cvss_vector)
        finding.cvss_score = float(data.get('cvss_score', finding.cvss_score))
        finding.severity = data.get('severity', finding.severity)
        finding.status = data.get('status', finding.status)
        finding.finding_status = data.get('finding_status', finding.finding_status)
        finding.retest_evidence = data.get('retest_evidence', finding.retest_evidence)
        finding.custom_fields = data.get('custom_fields', finding.custom_fields)
        db.session.commit()
        log_audit('UPDATE_FINDING', f"Updated finding: {finding.title} (Status: {finding.finding_status}, Retest Status: {finding.status})")
        return jsonify(finding.to_dict())
    elif request.method == 'DELETE':
        find_title = finding.title
        db.session.delete(finding)
        db.session.commit()
        log_audit('DELETE_FINDING', f"Deleted finding: {find_title}")
        return jsonify({'message': 'Finding deleted successfully'})

# FindingTemplate CRUD
@app.route('/api/finding_templates', methods=['GET', 'POST'])
@login_required
def api_finding_templates():
    if request.method == 'GET':
        from sqlalchemy.orm import defer
        deferred_cols = (
            defer(FindingTemplate.description), defer(FindingTemplate.exploitation),
            defer(FindingTemplate.impact), defer(FindingTemplate.solution),
            defer(FindingTemplate.reference), defer(FindingTemplate.step_reproduce),
            defer(FindingTemplate.custom_fields)
        )
        templates = FindingTemplate.query.options(*deferred_cols).order_by(FindingTemplate.cvss_score.desc()).all()
        return jsonify([t.to_dict() for t in templates])
    elif request.method == 'POST':
        data = request.json
        template = FindingTemplate(
            title=data['title'],
            description=data.get('description', ''),
            exploitation=data.get('exploitation', ''),
            impact=data.get('impact', ''),
            solution=data.get('solution', ''),
            reference=data.get('reference', ''),
            step_reproduce=data.get('step_reproduce', ''),
            cwe=data.get('cwe', ''),
            mitre_attack=data.get('mitre_attack', ''),
            iso_27001=data.get('iso_27001', ''),
            nist_control=data.get('nist_control', ''),
            ptes_phase=data.get('ptes_phase', ''),
            cvss_version=data.get('cvss_version', 'v3.1'),
            cvss_vector=data.get('cvss_vector', ''),
            cvss_score=float(data.get('cvss_score', 0.0)),
            severity=data.get('severity', 'Info'),
            custom_fields=data.get('custom_fields', '')
        )
        db.session.add(template)
        db.session.commit()
        log_audit('CREATE_FINDING_TEMPLATE', f"Created finding template: {template.title}")
        return jsonify(template.to_dict()), 201

@app.route('/api/finding_templates/<int:template_id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def api_finding_template(template_id):
    template = FindingTemplate.query.get_or_404(template_id)
    if request.method == 'GET':
        return jsonify(template.to_dict())
    elif request.method == 'PUT':
        data = request.json
        template.title = data['title']
        template.description = data.get('description', template.description)
        template.exploitation = data.get('exploitation', template.exploitation)
        template.impact = data.get('impact', template.impact)
        template.solution = data.get('solution', template.solution)
        template.reference = data.get('reference', template.reference)
        template.step_reproduce = data.get('step_reproduce', template.step_reproduce)
        template.cwe = data.get('cwe', template.cwe)
        template.mitre_attack = data.get('mitre_attack', template.mitre_attack)
        template.iso_27001 = data.get('iso_27001', template.iso_27001)
        template.nist_control = data.get('nist_control', template.nist_control)
        template.ptes_phase = data.get('ptes_phase', template.ptes_phase)
        template.cvss_version = data.get('cvss_version', template.cvss_version)
        template.cvss_vector = data.get('cvss_vector', template.cvss_vector)
        template.cvss_score = float(data.get('cvss_score', template.cvss_score))
        template.severity = data.get('severity', template.severity)
        template.custom_fields = data.get('custom_fields', template.custom_fields)
        db.session.commit()
        log_audit('UPDATE_FINDING_TEMPLATE', f"Updated finding template: {template.title}")
        return jsonify(template.to_dict())
    elif request.method == 'DELETE':
        temp_title = template.title
        db.session.delete(template)
        db.session.commit()
        log_audit('DELETE_FINDING_TEMPLATE', f"Deleted finding template: {temp_title}")
        return jsonify({'message': 'Finding template deleted successfully'})

# Consultant CRUD
@app.route('/api/consultants', methods=['GET', 'POST'])
@login_required
def api_consultants():
    user = User.query.get(session['user_id'])
    if request.method == 'GET':
        consultants = Consultant.query.order_by(Consultant.name).all()
        return jsonify([c.to_dict() for c in consultants])
    elif request.method == 'POST':
        if user.role != 'Admin':
            return jsonify({'message': 'Only Admins can add team members'}), 403
        data = request.json
        consultant = Consultant(
            name=data['name'],
            role=data.get('role', ''),
            email=data.get('email', '')
        )
        db.session.add(consultant)
        db.session.commit()
        log_audit('CREATE_CONSULTANT', f"Created consultant: {consultant.name}")
        return jsonify(consultant.to_dict()), 201

@app.route('/api/consultants/<int:consultant_id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def api_consultant(consultant_id):
    user = User.query.get(session['user_id'])
    consultant = Consultant.query.get_or_404(consultant_id)
    if request.method == 'GET':
        data = consultant.to_dict()
        projects = Project.query.filter(
            (Project.pentest_consultant_id == consultant_id) |
            (Project.retest_consultant_id == consultant_id)
        ).all()
        data['projects'] = []
        for p in projects:
            role = "Pentester" if p.pentest_consultant_id == consultant_id else "Retester"
            if p.pentest_consultant_id == consultant_id and p.retest_consultant_id == consultant_id:
                role = "Pentester & Retester"
            p_dict = p.to_dict()
            p_dict['assigned_role'] = role
            data['projects'].append(p_dict)
        return jsonify(data)
    elif request.method == 'PUT':
        if user.role != 'Admin':
            return jsonify({'message': 'Only Admins can modify team members'}), 403
        data = request.json
        consultant.name = data['name']
        consultant.role = data.get('role', consultant.role)
        consultant.email = data.get('email', consultant.email)
        db.session.commit()
        log_audit('UPDATE_CONSULTANT', f"Updated consultant: {consultant.name}")
        return jsonify(consultant.to_dict())
    elif request.method == 'DELETE':
        if user.role != 'Admin':
            return jsonify({'message': 'Only Admins can delete team members'}), 403
        cons_name = consultant.name
        db.session.delete(consultant)
        db.session.commit()
        log_audit('DELETE_CONSULTANT', f"Deleted consultant: {cons_name}")
        return jsonify({'message': 'Consultant deleted successfully'})

# --- Auth & User Management APIs ---

@app.route('/api/auth/login', methods=['POST'])
def api_login():
    data = request.json
    username = data.get('username')
    password = data.get('password')
    ip_address = request.headers.get('X-Forwarded-For', request.remote_addr)
    
    if not username or not password:
        return jsonify({'message': 'Username and password are required'}), 400
        
    # Check if IP is currently rate-limited/locked
    if ip_address in login_failures:
        lock_until = login_failures[ip_address].get('locked_until')
        if lock_until and datetime.utcnow() < lock_until:
            return jsonify({'message': 'Too many failed login attempts. Your IP has been temporarily locked.'}), 429

    user = User.query.filter_by(username=username).first()
    if user and check_password_hash(user.password_hash, password):
        # Check if disabled
        is_disabled = user.is_disabled
        if not user.mfa_enabled and user.created_at:
            if (datetime.utcnow() - user.created_at) > timedelta(hours=24):
                user.is_disabled = True
                db.session.commit()
                is_disabled = True
        
        if is_disabled:
            log_audit('LOGIN_DISABLED', f'Failed login attempt for disabled user {username}')
            return jsonify({'message': 'Akun dinonaktifkan karena tidak mengaktifkan MFA dalam waktu 24 jam. Hubungi Administrator.'}), 403

        # Check if MFA is enabled
        if user.mfa_enabled:
            token = data.get('token')
            if not token:
                return jsonify({'status': 'mfa_required', 'username': user.username})
            
            # Verify TOTP token
            totp = pyotp.TOTP(user.mfa_secret)
            if not totp.verify(token):
                log_audit('LOGIN_MFA_FAILED', f'Invalid MFA token for user {username}')
                return jsonify({'message': 'Token MFA tidak valid.'}), 401

        session.clear()
        session['user_id'] = user.id
        session['username'] = user.username
        session['role'] = user.role
        
        # Reset failures on success
        if ip_address in login_failures:
            del login_failures[ip_address]
            
        log_audit('LOGIN', f'User {username} logged in successfully')
        return jsonify(user.to_dict())
    
    # Increment failures on failure
    if ip_address not in login_failures:
        login_failures[ip_address] = {'attempts': 0, 'locked_until': None}
    
    login_failures[ip_address]['attempts'] += 1
    attempts = login_failures[ip_address]['attempts']
    
    if attempts >= 5:
        lock_time = datetime.utcnow() + timedelta(minutes=15)
        login_failures[ip_address]['locked_until'] = lock_time
        
        # Also record in BlockedIP database for persistent block
        block_entry = BlockedIP.query.filter_by(ip_address=ip_address).first()
        if not block_entry:
            block_entry = BlockedIP(
                ip_address=ip_address,
                reason="Brute-force Login Protection (5 failed attempts)",
                expires_at=lock_time
            )
            db.session.add(block_entry)
            
        log_audit('BRUTE_FORCE_LOCK', f'IP {ip_address} locked out for 15 mins due to 5 login failures.')
        db.session.commit()
        return jsonify({'message': 'Too many failed login attempts. Your IP has been temporarily blocked for 15 minutes.'}), 429
    
    # Audit log failed login
    log_entry = AuditLog(
        action='LOGIN_FAILED',
        details=f'Failed login attempt for username: {username} (Attempt {attempts}/5)',
        ip_address=ip_address
    )
    db.session.add(log_entry)
    db.session.commit()
    return jsonify({'message': f'Invalid username or password. Attempt {attempts}/5'}), 401

@app.route('/api/user/mfa/setup', methods=['GET', 'POST'])
@login_required
def api_user_mfa_setup():
    user = User.query.get(session['user_id'])
    if request.method == 'GET':
        if not user.mfa_secret:
            user.mfa_secret = pyotp.random_base32()
            db.session.commit()
        totp = pyotp.TOTP(user.mfa_secret)
        otpauth_url = totp.provisioning_uri(name=user.username, issuer_name="PentaGO")
        return jsonify({
            'secret': user.mfa_secret,
            'otpauth_url': otpauth_url,
            'mfa_enabled': user.mfa_enabled
        })
    elif request.method == 'POST':
        data = request.json
        token = data.get('token')
        if not token:
            return jsonify({'message': 'Verification token required'}), 400
        if not user.mfa_secret:
            return jsonify({'message': 'MFA setup not initialized'}), 400
        totp = pyotp.TOTP(user.mfa_secret)
        if totp.verify(token):
            user.mfa_enabled = True
            db.session.commit()
            log_audit('MFA_ENABLED', f"MFA enabled for user {user.username}")
            return jsonify({'message': 'MFA successfully enabled', 'mfa_enabled': True})
        else:
            return jsonify({'message': 'Invalid verification code'}), 400

@app.route('/api/user/mfa/disable', methods=['POST'])
@login_required
def api_user_mfa_disable():
    user = User.query.get(session['user_id'])
    data = request.json
    token = data.get('token')
    if not token:
        return jsonify({'message': 'Verification token required'}), 400
    if not user.mfa_secret:
        return jsonify({'message': 'MFA not configured'}), 400
    totp = pyotp.TOTP(user.mfa_secret)
    if totp.verify(token):
        user.mfa_enabled = False
        user.mfa_secret = None
        db.session.commit()
        log_audit('MFA_DISABLED', f"MFA disabled for user {user.username}")
        return jsonify({'message': 'MFA successfully disabled', 'mfa_enabled': False})
    else:
        return jsonify({'message': 'Invalid verification code'}), 400

@app.route('/api/auth/logout', methods=['POST'])
@login_required
def api_logout():
    username = session.get('username')
    log_audit('LOGOUT', f'User {username} logged out')
    session.clear()
    return jsonify({'message': 'Logged out successfully'})

@app.route('/api/auth/me', methods=['GET'])
def api_me():
    if 'user_id' not in session:
        return jsonify({'logged_in': False}), 200
    user = User.query.get(session['user_id'])
    if not user:
        session.clear()
        return jsonify({'logged_in': False}), 200
        
    # Check if they should be disabled now
    if not user.mfa_enabled and user.created_at:
        if (datetime.utcnow() - user.created_at) > timedelta(hours=24):
            user.is_disabled = True
            db.session.commit()
            log_audit('LOGIN_DISABLED', f'User {user.username} disabled due to missing MFA after 24h limit')
            session.clear()
            return jsonify({'logged_in': False, 'message': 'Akun dinonaktifkan karena tidak mengaktifkan MFA dalam waktu 24 jam.'}), 200

    res = user.to_dict()
    res['logged_in'] = True
    return jsonify(res)

@app.route('/api/auth/change-password', methods=['POST'])
@login_required
def api_change_password():
    data = request.json
    old_password = data.get('old_password')
    new_password = data.get('new_password')
    
    if not old_password or not new_password:
        return jsonify({'message': 'Old and new passwords are required'}), 400
        
    user = User.query.get(session['user_id'])
    if not check_password_hash(user.password_hash, old_password):
        return jsonify({'message': 'Incorrect old password'}), 400
        
    user.password_hash = generate_password_hash(new_password)
    db.session.commit()
    log_audit('CHANGE_PASSWORD', f'User {user.username} changed their own password')
    return jsonify({'message': 'Password updated successfully'})

# --- Admin APIs ---

@app.route('/api/admin/users', methods=['GET', 'POST'])
@admin_required
def api_admin_users():
    if request.method == 'GET':
        users = User.query.order_by(User.username).all()
        return jsonify([u.to_dict() for u in users])
    elif request.method == 'POST':
        data = request.json
        username = data.get('username')
        password = data.get('password')
        role = data.get('role', 'User')
        allowed_companies = data.get('allowed_companies', [])
        
        if not username or not password:
            return jsonify({'message': 'Username and password are required'}), 400
            
        if User.query.filter_by(username=username).first():
            return jsonify({'message': 'Username already exists'}), 400
            
        fullname = data.get('fullname') or username
        email = data.get('email', '')
        member_role = data.get('member_role', 'Consultant')
        
        # Automatically create Consultant profile
        consultant = Consultant(name=fullname, role=member_role, email=email)
        db.session.add(consultant)
        db.session.flush() # get consultant.id
        
        new_user = User(
            username=username,
            password_hash=generate_password_hash(password),
            role=role,
            consultant_id=consultant.id
        )
        if allowed_companies:
            companies = Company.query.filter(Company.id.in_(allowed_companies)).all()
            new_user.allowed_companies = companies
            
        db.session.add(new_user)
        db.session.commit()
        log_audit('CREATE_USER', f'Created user {username} with role {role}')
        return jsonify(new_user.to_dict()), 201

@app.route('/api/admin/users/<int:user_id>', methods=['PUT', 'DELETE'])
@admin_required
def api_admin_user_detail(user_id):
    user = User.query.get_or_404(user_id)
    
    if request.method == 'PUT':
        data = request.json
        new_role = data.get('role')
        new_password = data.get('password')
        allowed_companies = data.get('allowed_companies')
        
        if new_password:
            user.password_hash = generate_password_hash(new_password)
            log_audit('RESET_PASSWORD', f'Reset password for user {user.username}')
            
        if new_role and new_role in ['Admin', 'User']:
            if user.id == session['user_id'] and new_role != 'Admin':
                return jsonify({'message': 'Cannot demote your own admin account'}), 400
            old_role = user.role
            user.role = new_role
            log_audit('UPDATE_USER_ROLE', f'Changed role of {user.username} from {old_role} to {new_role}')
            
        fullname = data.get('fullname')
        email = data.get('email')
        member_role = data.get('member_role')
        
        if user.consultant:
            if fullname:
                user.consultant.name = fullname
            if email is not None:
                user.consultant.email = email
            if member_role:
                user.consultant.role = member_role
        else:
            consultant = Consultant(name=fullname or user.username, role=member_role or 'Consultant', email=email or '')
            db.session.add(consultant)
            db.session.flush()
            user.consultant_id = consultant.id

        if allowed_companies is not None:
            companies = Company.query.filter(Company.id.in_(allowed_companies)).all()
            user.allowed_companies = companies
            
        is_disabled = data.get('is_disabled')
        reset_mfa = data.get('reset_mfa')
        
        if is_disabled is not None:
            user.is_disabled = bool(is_disabled)
            if not is_disabled:
                # Giving them a fresh 24 hours to enable MFA
                user.created_at = datetime.utcnow()
            log_audit('UPDATE_USER_STATUS', f"{'Disabled' if is_disabled else 'Enabled'} user {user.username}")
            
        if reset_mfa:
            user.mfa_enabled = False
            user.mfa_secret = None
            # Giving them a fresh 24 hours to setup MFA
            user.created_at = datetime.utcnow()
            log_audit('RESET_MFA', f"Reset MFA for user {user.username}")

        db.session.commit()
        return jsonify(user.to_dict())
        
    elif request.method == 'DELETE':
        # Prevent self-deletion
        if user.id == session['user_id']:
            return jsonify({'message': 'Cannot delete your own admin account'}), 400
            
        username = user.username
        if user.consultant:
            db.session.delete(user.consultant)
        db.session.delete(user)
        db.session.commit()
        log_audit('DELETE_USER', f'Deleted user account: {username}')
        return jsonify({'message': 'User deleted successfully'})

@app.route('/api/admin/audit-logs', methods=['GET', 'DELETE'])
@admin_required
def api_admin_audit_logs():
    if request.method == 'GET':
        query = AuditLog.query
        
        # Filter by search query
        search = request.args.get('search')
        if search:
            query = query.filter(
                (AuditLog.username.like(f'%{search}%')) |
                (AuditLog.action.like(f'%{search}%')) |
                (AuditLog.details.like(f'%{search}%'))
            )
            
        logs = query.order_by(AuditLog.timestamp.desc()).limit(200).all()
        return jsonify([l.to_dict() for l in logs])
    elif request.method == 'DELETE':
        AuditLog.query.delete()
        db.session.commit()
        log_audit('CLEAR_LOGS', f'Audit logs were cleared by {session.get("username")}')
        return jsonify({'message': 'Audit logs cleared successfully'})

@app.route('/api/admin/blocklist', methods=['GET', 'POST'])
@admin_required
def api_admin_blocklist():
    if request.method == 'GET':
        blocks = BlockedIP.query.order_by(BlockedIP.blocked_at.desc()).all()
        return jsonify([b.to_dict() for b in blocks])
    elif request.method == 'POST':
        data = request.json
        ip = data.get('ip_address')
        reason = data.get('reason', 'Manually blocked by Admin')
        expires_in_mins = data.get('expires_in') # None for permanent, otherwise integer minutes
        
        if not ip:
            return jsonify({'message': 'IP address is required'}), 400
            
        expires_at = None
        if expires_in_mins:
            expires_at = datetime.utcnow() + timedelta(minutes=int(expires_in_mins))
            
        existing = BlockedIP.query.filter_by(ip_address=ip).first()
        if existing:
            existing.reason = reason
            existing.expires_at = expires_at
            block_entry = existing
            log_audit('UPDATE_BLOCKLIST', f'Updated blocklist settings for IP: {ip}')
        else:
            block_entry = BlockedIP(ip_address=ip, reason=reason, expires_at=expires_at)
            db.session.add(block_entry)
            log_audit('BLOCK_IP', f'Blocked IP address: {ip} (Reason: {reason})')
            
        db.session.commit()
        return jsonify(block_entry.to_dict()), 201

@app.route('/api/admin/blocklist/<int:block_id>', methods=['DELETE'])
@admin_required
def api_admin_blocklist_delete(block_id):
    block = BlockedIP.query.get_or_404(block_id)
    ip = block.ip_address
    db.session.delete(block)
    db.session.commit()
    
    # Also release in-memory login failures lock if any
    if ip in login_failures:
        del login_failures[ip]
        
    log_audit('UNBLOCK_IP', f'Unblocked IP address: {ip}')
    return jsonify({'message': f'IP {ip} unblocked successfully'})

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        # Automatically unblock localhost if it got blocked previously
        BlockedIP.query.filter(BlockedIP.ip_address.in_(['127.0.0.1', '::1', 'localhost'])).delete(synchronize_session=False)
        db.session.commit()
    app.run(debug=True, port=5001)
